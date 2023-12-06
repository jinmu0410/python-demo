import os
import urllib
import requests
import json
import pandas as pd
from sqlalchemy import create_engine

import time
from openpyxl import load_workbook
import re


def query(a1, a2, a3, a4, a5):
    headers = {
        'Cookie':'UM_distinctid=18be54b16891124-030f6b11141f9d-16525634-1fa400-18be54b168a1a58; Hm_lvt_6b63cf9e50e2bd684eba62e24995ba09=1700358985; Hm_lvt_78d5885b19eecf93e59673b4b37c8530=1700358985; cna=2a89576bbc8f4def90f4c035046ccf36; tenantId=TPN1726055943878574082; sysAccountId=202311190100010003UEVR0000000082; Hm_lpvt_78d5885b19eecf93e59673b4b37c8530=1700371324; Hm_lpvt_6b63cf9e50e2bd684eba62e24995ba09=1700371324; acw_tc=76b20f6817004430272062566e762e4440bd81e2652337cbf7bd76abce581a; X-REALM-ORG=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJjdHgiOiJKU0VtNytteXZhcktSTVZTWnBMRzJ4RS9hVXc1VXlWVkVWYjgwbWc0U2RuSUFGcVFBKzk2NmVpR2NyemoydmhMaTNlTm1vbEJaeXJFSk53UEVxdktuWjRLTS9KcHZKL3dLTEpaWjFDS0FneE5qWVF6dGVRUXJRZ1ZYNmdmZ1YyNWJteDJITFd5ejJNSXBvUDY4RXkxQzc0K0d1WHlKVjlGdlJhdkV6SEJRa1U1VTg0V1FUTDRKMEJPbXBDNjVzVDNXSFlRZGVpMDlvYTZiWEU1UEpMMzEzNEFpdHBlcEZZU3gvOW53Tlc2ZkVseDhHK1k2YzFaVm55Y01NTXhZSE1SQ2VUcjc0MzJFNmVOZTYrTVZZelBVd1V2WUFaS1g2OEZ3QTh5dFFCSk40N0NWY00xc2k0aVJ4cC9zSkkwUG5RemdIQnpPVUdGaDArMjdxeVp3K2JqWHNYN3Vmamp6T3R5N09IcE5raTdQUkRyQ0ZMRU9rRzIzMDBVRnpCeVZWbkxySW4yOFprTFIzTFJnTWE4RTV0bSs5MHRrYTdVOWV3czFnSXhhckJnYXIwPSIsImlzcyI6IlFKRCIsImVuYyI6dHJ1ZSwiZXhwIjoxNzAxMDQ3ODI3LCJpYXQiOjE3MDA0NDMwMjcsImp0aSI6ImFjZjU0YWUyLWMyODEtNDI0My1iZDJlLWY3NjdiYzdiM2RkMCJ9.YMY0ho-wvdiTKM4fi2Oo-15KTD7PHEXIQFwv2aHkO0Y',
        'Origin': 'https://www.sscha.com',
        'Referer': 'https://www.sscha.com/',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'sec-ch-ua': 'Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Content-Type': 'application/json;charset=UTF-8',
        'Host': 'api.sscha.com'
    }

    str = '{"countryCodeList":[args1],"status":[1],"companyType":[args2],"searchType":21,"pageIndex":args3,"pageSize":args4}'

    data = str.replace('args1', a1, 1).replace('args2', a2, 1).replace('args3', a3, 1).replace('args4', a4, 1)

    response = requests.post("https://api.sscha.com/center/company/pageCompany", headers=headers, data=data)

    if (response.status_code == 200):
        res = json.loads(response.text)['data']
        if res is None:
            print(response.text)
            return
        else:
            print("准备写入文件中")
            # saveExcel(json.loads(response.text)['data'], a5.split('-', 1)[0] + '.xlsx', a5)
            save_to_mysql(json.loads(response.text)['data'])


def data_clean(text):
    # 清洗excel中的非法字符，都是不常见的不可显示字符，例如退格，响铃等
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
    return text

def check_sheet_exists(file_name, sheet_name):
    try:
        workbook = load_workbook(file_name, read_only=True)
        sheet_names = workbook.sheetnames
        return sheet_name in sheet_names
    except FileNotFoundError:
        return False

def saveExcel(data, file_name, sheet_name):
    df = pd.DataFrame(data)
    df = df.fillna('').astype(str)
    for col in df.columns:
        df[col] = df[col].apply(lambda x: data_clean(x))

    if os.path.isfile(file_name):
        # existing_data = pd.read_excel(file_name, sheet_name=None)
        sheet_exists=check_sheet_exists(file_name,sheet_name)
        if sheet_exists:
            existing_data = pd.read_excel(file_name, sheet_name=sheet_name)
            start_row = existing_data.shape[0] + 1
            print('开始写入1')
            with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=start_row)
        else:
            print('开始写入2')
            with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, sheet_name=sheet_name)
    else:
        df.to_excel(file_name, sheet_name=sheet_name, index=False)
    print("保存成功:"+file_name)

def save_to_mysql(data):
    # 创建一个连接到 MySQL 数据库的引擎
    # 格式为：'mysql+mysqlconnector://username:password@hostname:port/database_name'
    encoded_password = urllib.parse.quote('R#m9@L7s$yFhN23')

    connection_string = f'mysql+mysqlconnector://aifish:{encoded_password}@47.97.231.182:3306/spider_man'
    engine = create_engine(connection_string)

    # 假设 df 是你的 Pandas DataFrame
    df = pd.DataFrame(data)
    df = df.astype(str)
    # 显式地打开连接
    connection = engine.connect()
    try:
        # 将 DataFrame 写入 MySQL 数据库中的表格
        table_name = 'company_data'  # 设定表格名称
        df.to_sql(table_name, con=connection, if_exists='append', index=False)

        # 数据写入完成后，手动关闭连接
        connection.close()
    except Exception as e:
        # 发生异常时关闭连接并处理异常
        connection.close()
        print("发生错误:", e)

# 读取本地JSON文件
file_path = 'yes/广西.json'  # 替换成你的JSON文件路径

with open(file_path, 'r', encoding='utf-8') as file:
    data = json.load(file)


# 获取最后一层级的extName并进行拼接
def get_final_ext_names(data, prefix=''):
    result = []
    if isinstance(data, dict):
        if 'extName' in data:
            prefix = prefix + data['extName'] + '-' if prefix else data['extName'] + '-'
            if 'children' not in data:
                result.append((data['id'], prefix[:-1]))
        for value in data.values():
            if isinstance(value, (dict, list)):
                result.extend(get_final_ext_names(value, prefix))
    elif isinstance(data, list):
        for item in data:
            result.extend(get_final_ext_names(item, prefix))
    else:
        return [(prefix[:-1])] if prefix else []
    return result


if __name__ == '__main__':

    for element in data:
        # 构建id和extName的字典
        id_ext_name_map = dict(get_final_ext_names(element))

        for id, ext_name in id_ext_name_map.items():
            print(f"{id}: {ext_name}")
            print(ext_name.split('-', 1)[0])
            param1 = 1
            while param1 <= 5:  # 循环直到参数2达到1W
                query(id, '5', str(param1), '1000',ext_name)
                param1 += 1
                print(param1)
                time.sleep(1.5)
            #time.sleep(1)

#
# if __name__ == '__main__':
#     param1 = 1
#     while param1 <= 5:  # 循环直到参数2达到1W
#         query('440106', '5', str(param1), '1000', '广东省-广州市-天河区')
#         param1 += 1
#         print(param1)
#         time.sleep(4)